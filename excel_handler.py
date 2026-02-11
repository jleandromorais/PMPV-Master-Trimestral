# -*- coding: utf-8 -*-
"""
Módulo de Integração com Excel para a Calculadora PMPV
Permite importar e exportar dados em formato Excel
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List
import os


class ExcelHandlerPMPV:
    """Gerenciador de importação/exportação Excel"""
    
    @staticmethod
    def exportar_trimestre(dados_por_mes: Dict, resultado: Dict, 
                          nome_arquivo: str = None) -> str:
        """
        Exporta dados do trimestre para Excel formatado.
        
        Args:
            dados_por_mes: Dicionário com dados dos 3 meses
            resultado: Dicionário com resultado do cálculo
            nome_arquivo: Nome do arquivo (se None, usa timestamp)
            
        Returns:
            Caminho do arquivo criado
        """
        if nome_arquivo is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_arquivo = f"PMPV_Trimestral_{timestamp}.xlsx"
        
        # Criar workbook
        wb = openpyxl.Workbook()
        
        # Remover sheet padrão
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Criar abas para cada mês
        for i in range(1, 4):
            mes_nome = f"Mês {i}"
            dados_mes = dados_por_mes.get(mes_nome, [])
            ExcelHandlerPMPV._criar_aba_mes(wb, mes_nome, dados_mes, i)
        
        # Criar aba de resumo
        ExcelHandlerPMPV._criar_aba_resumo(wb, dados_por_mes, resultado)
        
        # Salvar
        wb.save(nome_arquivo)
        print(f"Excel exportado: {nome_arquivo}")
        return nome_arquivo
    
    @staticmethod
    def _criar_aba_mes(wb, nome_aba: str, dados: List[Dict], mes_num: int):
        """Cria uma aba para um mês específico"""
        ws = wb.create_sheet(nome_aba)
        
        # Estilos
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:G1')
        ws['A1'] = f"PMPV - {nome_aba}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Cabeçalhos
        headers = ['Empresa/Contrato', 'Molécula (R$/m³)', 'Transporte (R$/m³)', 
                  'Logística (R$/m³)', 'Preço Final (R$/m³)', 'Volume (m³/dia)', 'Custo Total (R$)']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Dados
        row_num = 4
        for linha in dados:
            if not linha.get('volume', 0):  # Pula linhas vazias
                continue
            
            empresa = linha.get('empresa', '')
            mol = float(linha.get('molecula', 0))
            trans = float(linha.get('transporte', 0))
            log = float(linha.get('logistica', 0))
            vol = float(linha.get('volume', 0))
            
            preco_final = mol + trans + log
            custo_total = preco_final * vol
            
            ws.cell(row=row_num, column=1).value = empresa
            ws.cell(row=row_num, column=2).value = mol
            ws.cell(row=row_num, column=2).number_format = '#,##0.0000'
            ws.cell(row=row_num, column=3).value = trans
            ws.cell(row=row_num, column=3).number_format = '#,##0.0000'
            ws.cell(row=row_num, column=4).value = log
            ws.cell(row=row_num, column=4).number_format = '#,##0.0000'
            ws.cell(row=row_num, column=5).value = preco_final
            ws.cell(row=row_num, column=5).number_format = '#,##0.0000'
            ws.cell(row=row_num, column=6).value = vol
            ws.cell(row=row_num, column=6).number_format = '#,##0.00'
            ws.cell(row=row_num, column=7).value = custo_total
            ws.cell(row=row_num, column=7).number_format = '#,##0.00'
            
            # Bordas
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).border = border
            
            row_num += 1
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 20
    
    @staticmethod
    def _criar_aba_resumo(wb, dados_por_mes: Dict, resultado: Dict):
        """Cria aba de resumo trimestral"""
        ws = wb.create_sheet("Resumo Trimestral", 0)  # Primeira aba
        
        # Título
        ws.merge_cells('A1:D1')
        ws['A1'] = "RESUMO TRIMESTRAL - PMPV"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Data
        ws['A2'] = f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(italic=True)
        
        # Resultados
        row = 4
        ws[f'A{row}'] = "RESULTADOS DO TRIMESTRE"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="2C3E50")
        
        row += 2
        ws[f'A{row}'] = "Volume Total Acumulado:"
        ws[f'B{row}'] = resultado.get('volume_total', 0)
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'C{row}'] = "m³"
        
        row += 1
        ws[f'A{row}'] = "Custo Total Estimado:"
        ws[f'B{row}'] = resultado.get('custo_total', 0)
        ws[f'B{row}'].number_format = 'R$ #,##0.00'
        
        row += 1
        ws[f'A{row}'] = "PMPV Trimestral:"
        ws[f'B{row}'] = resultado.get('pmpv', 0)
        ws[f'B{row}'].number_format = 'R$ #,##0.0000'
        ws[f'C{row}'] = "/m³"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'B{row}'].font = Font(bold=True, size=11, color="27AE60")
        
        # Breakdown por mês
        row += 3
        ws[f'A{row}'] = "DETALHAMENTO POR MÊS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="2C3E50")
        
        row += 2
        # Cabeçalho
        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        for col, header in enumerate(['Mês', 'Volume (m³)', 'Custo (R$)', 'PMPV (R$/m³)'], start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        for i in range(1, 4):
            mes_nome = f"Mês {i}"
            dados_mes = dados_por_mes.get(mes_nome, [])
            
            volume_mes = 0.0
            custo_mes = 0.0
            
            for linha in dados_mes:
                vol = float(linha.get('volume', 0))
                if vol > 0:
                    mol = float(linha.get('molecula', 0))
                    trans = float(linha.get('transporte', 0))
                    log = float(linha.get('logistica', 0))
                    preco = mol + trans + log
                    
                    volume_mes += vol
                    custo_mes += preco * vol
            
            pmpv_mes = custo_mes / volume_mes if volume_mes > 0 else 0
            
            ws.cell(row=row, column=1).value = mes_nome
            ws.cell(row=row, column=2).value = volume_mes
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=3).value = custo_mes
            ws.cell(row=row, column=3).number_format = 'R$ #,##0.00'
            ws.cell(row=row, column=4).value = pmpv_mes
            ws.cell(row=row, column=4).number_format = 'R$ #,##0.0000'
            
            row += 1
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
    
    @staticmethod
    def importar_excel(caminho_arquivo: str) -> Dict:
        """
        Importa dados de um arquivo Excel.
        
        Args:
            caminho_arquivo: Caminho do arquivo Excel
            
        Returns:
            Dicionário com dados importados
        """
        if not os.path.exists(caminho_arquivo):
            raise FileNotFoundError(f"Arquivo não encontrado: {caminho_arquivo}")
        
        wb = openpyxl.load_workbook(caminho_arquivo)
        dados_por_mes = {}
        
        # Procurar abas de meses
        for i in range(1, 4):
            mes_nome = f"Mês {i}"
            if mes_nome in wb.sheetnames:
                ws = wb[mes_nome]
                dados_mes = []
                
                # Começar da linha 4 (depois do cabeçalho)
                for row in range(4, ws.max_row + 1):
                    empresa = ws.cell(row=row, column=1).value
                    if not empresa:
                        continue
                    
                    dados_mes.append({
                        'empresa': empresa,
                        'molecula': ws.cell(row=row, column=2).value or 0,
                        'transporte': ws.cell(row=row, column=3).value or 0,
                        'logistica': ws.cell(row=row, column=4).value or 0,
                        'volume': ws.cell(row=row, column=6).value or 0
                    })
                
                dados_por_mes[mes_nome] = dados_mes
        
        wb.close()
        return dados_por_mes
    
    @staticmethod
    def criar_template(nome_arquivo: str = "PMPV_Template.xlsx"):
        """
        Cria um arquivo Excel template vazio para preenchimento manual.
        
        Args:
            nome_arquivo: Nome do arquivo template
            
        Returns:
            Caminho do arquivo criado
        """
        # Dados vazios
        dados_vazios = {
            'Mês 1': [],
            'Mês 2': [],
            'Mês 3': []
        }
        
        resultado_vazio = {
            'volume_total': 0,
            'custo_total': 0,
            'pmpv': 0
        }
        
        return ExcelHandlerPMPV.exportar_trimestre(
            dados_vazios, 
            resultado_vazio, 
            nome_arquivo
        )


# Exemplo de uso
if __name__ == "__main__":
    # Teste de exportação
    dados_teste = {
        'Mês 1': [
            {'empresa': 'Fornecedor 1', 'molecula': 10.50, 'transporte': 0.50, 'logistica': 0.30, 'volume': 100000},
            {'empresa': 'Fornecedor 2', 'molecula': 11.20, 'transporte': 0.45, 'logistica': 0.25, 'volume': 80000},
        ],
        'Mês 2': [
            {'empresa': 'Fornecedor 1', 'molecula': 10.50, 'transporte': 0.50, 'logistica': 0.30, 'volume': 100000},
        ],
        'Mês 3': []
    }
    
    resultado_teste = {
        'volume_total': 280000,
        'custo_total': 3196000,
        'pmpv': 11.41
    }
    
    arquivo = ExcelHandlerPMPV.exportar_trimestre(dados_teste, resultado_teste, "teste_pmpv.xlsx")
    print(f"Teste concluído! Arquivo: {arquivo}")
    
    # Criar template
    template = ExcelHandlerPMPV.criar_template()
    print(f"Template criado: {template}")
